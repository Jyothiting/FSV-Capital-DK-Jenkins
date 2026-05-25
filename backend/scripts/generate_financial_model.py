"""
Generate FSV Capital investor-grade financial model workbooks.

Outputs:
  docs/financial-model/FSV_Financial_Model_Template.xlsx  (blank template)
  docs/financial-model/samples/Apex_AI_Labs_Financial_Model.xlsx  (filled demo)

Run from repo root:
  .venv\\Scripts\\python backend/scripts/generate_financial_model.py
"""
from __future__ import annotations

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
OUT_DIR = os.path.join(ROOT, "docs", "financial-model")
SAMPLES_DIR = os.path.join(OUT_DIR, "samples")

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="E8F0FE")
SECTION_FONT = Font(bold=True, size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
USD_FMT = '"$"#,##0'
PCT_FMT = "0.0%"


def _style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _style_section(ws, row: int, text: str, cols: int = 6) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT


def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _add_assumptions(ws, data: dict) -> None:
    ws.title = "Assumptions"
    _set_col_widths(ws, {1: 32, 2: 18, 3: 40})

    ws["A1"] = "FSV Capital — Startup Financial Model"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Maps to funding form: Step 6 Financials, Step 7 Funding, Step 10 Documents"
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    rows = [
        ("", ""),
        ("COMPANY PROFILE", ""),
        ("Startup name", data.get("startup_name", "")),
        ("Industry / sector", data.get("industry_sector", "AI / ML")),
        ("Business model", data.get("business_model", "SaaS")),
        ("Current stage", data.get("current_stage", "MVP")),
        ("HQ / currency", data.get("currency", "USD")),
        ("Model date", data.get("model_date", str(date.today()))),
        ("", ""),
        ("TRACTION INPUTS (Step 5)", ""),
        ("Starting MRR (USD)", data.get("starting_mrr", 8000)),
        ("MoM revenue growth", data.get("mom_growth", 0.12)),
        ("Monthly logo churn", data.get("churn", 0.03)),
        ("Starting paying customers", data.get("starting_customers", 24)),
        ("ARPU / month (USD)", data.get("arpu", 333)),
        ("", ""),
        ("CAPITAL & BURN (Step 6)", ""),
        ("Funding raised to date (USD)", data.get("prior_raise", 150000)),
        ("Cash on hand (USD)", data.get("cash_on_hand", 180000)),
        ("Monthly burn today (USD)", data.get("current_burn", 42000)),
        ("", ""),
        ("THIS ROUND (Step 7)", ""),
        ("Amount raising (USD)", data.get("amount_raising", 500000)),
        ("Funding stage", data.get("funding_stage", "Seed")),
        ("Equity offered", data.get("equity_offered", "10%")),
        ("", ""),
        ("COST ASSUMPTIONS", ""),
        ("COGS % of revenue", data.get("cogs_pct", 0.35)),
        ("R&D % of OpEx", data.get("rd_pct", 0.55)),
        ("Sales & marketing % of OpEx", data.get("sm_pct", 0.30)),
        ("G&A % of OpEx", data.get("ga_pct", 0.15)),
        ("Headcount (FTE) — start", data.get("headcount_start", 6)),
        ("Headcount (FTE) — month 36", data.get("headcount_end", 14)),
        ("Avg fully-loaded cost / FTE / mo", data.get("fte_cost", 8500)),
    ]

    r = 4
    for label, value in rows:
        if label.startswith("COMPANY") or label.startswith("TRACTION") or label.startswith("CAPITAL") or label.startswith("THIS") or label.startswith("COST"):
            _style_section(ws, r, label, 3)
            r += 1
            continue
        if label == "" and value == "":
            r += 1
            continue
        ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=value)
        if isinstance(value, (int, float)) and (
            "growth" in label.lower() or "churn" in label.lower() or "%" in label
        ):
            c2.number_format = PCT_FMT
        elif isinstance(value, (int, float)):
            c2.number_format = USD_FMT
        r += 1


def _month_labels(n: int = 36) -> list[str]:
    return [f"M{m}" for m in range(1, n + 1)]


def _project_revenue(data: dict, months: int = 36) -> list[dict]:
    mrr = float(data.get("starting_mrr", 8000))
    growth = float(data.get("mom_growth", 0.12))
    churn = float(data.get("churn", 0.03))
    customers = int(data.get("starting_customers", 24))
    arpu = float(data.get("arpu", 333))
    rows = []
    for m in range(1, months + 1):
        if m > 1:
            mrr = mrr * (1 + growth) * (1 - churn)
            customers = max(customers, int(round(mrr / arpu)) if arpu else customers + 2)
        revenue = mrr
        cogs_pct = float(data.get("cogs_pct", 0.35))
        cogs = revenue * cogs_pct
        gross = revenue - cogs
        burn_base = float(data.get("current_burn", 42000))
        head_start = int(data.get("headcount_start", 6))
        head_end = int(data.get("headcount_end", 14))
        fte_cost = float(data.get("fte_cost", 8500))
        headcount = head_start + (head_end - head_start) * (m - 1) / (months - 1)
        payroll = headcount * fte_cost
        opex = max(burn_base * 0.65, payroll)
        rd = opex * float(data.get("rd_pct", 0.55))
        sm = opex * float(data.get("sm_pct", 0.30))
        ga = opex * float(data.get("ga_pct", 0.15))
        total_opex = rd + sm + ga
        ebitda = gross - total_opex
        net_burn = total_opex + cogs - revenue  # cash burn when negative ebitda
        if ebitda >= 0:
            net_burn = max(0, total_opex + cogs - revenue)
        else:
            net_burn = abs(ebitda)
        rows.append(
            {
                "month": m,
                "mrr": mrr,
                "arr": mrr * 12,
                "customers": customers,
                "revenue": revenue,
                "cogs": cogs,
                "gross_profit": gross,
                "rd": rd,
                "sm": sm,
                "ga": ga,
                "opex": total_opex,
                "ebitda": ebitda,
                "net_burn": net_burn,
            }
        )
    return rows


def _add_monthly_pl(ws, projections: list[dict]) -> None:
    ws.title = "Monthly P&L"
    headers = [
        "Month",
        "MRR",
        "ARR",
        "Customers",
        "Revenue",
        "COGS",
        "Gross Profit",
        "R&D",
        "S&M",
        "G&A",
        "Total OpEx",
        "EBITDA",
        "Net Cash Burn",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, row in enumerate(projections, 2):
        vals = [
            row["month"],
            row["mrr"],
            row["arr"],
            row["customers"],
            row["revenue"],
            row["cogs"],
            row["gross_profit"],
            row["rd"],
            row["sm"],
            row["ga"],
            row["opex"],
            row["ebitda"],
            row["net_burn"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = BORDER
            if c in (2, 3, 5, 6, 7, 8, 9, 10, 11, 12, 13):
                cell.number_format = USD_FMT
    _set_col_widths(ws, {1: 8, 2: 12, 3: 12, 4: 10})


def _add_cashflow(ws, projections: list[dict], data: dict) -> None:
    ws.title = "Cash Flow"
    cash = float(data.get("cash_on_hand", 180000))
    prior = float(data.get("prior_raise", 150000))
    new_round = float(data.get("amount_raising", 500000))
    round_month = int(data.get("round_close_month", 1))

    headers = ["Month", "Opening Cash", "Funding In", "Revenue In", "Cash Out (Burn)", "Closing Cash", "Runway (mo)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, p in enumerate(projections, 2):
        opening = cash
        funding = new_round if p["month"] == round_month else 0
        revenue_in = p["revenue"]
        cash_out = p["net_burn"] + p["cogs"] + p["opex"] - p["revenue"] if p["ebitda"] < 0 else p["net_burn"]
        if cash_out < 0:
            cash_out = p["net_burn"]
        closing = opening + funding + revenue_in - cash_out
        # runway: months until cash < 0 at current burn
        burn_rate = p["net_burn"] if p["net_burn"] > 0 else 1
        runway = closing / burn_rate if burn_rate else 99
        ws.cell(row=i, column=1, value=p["month"])
        for col, val in enumerate([opening, funding, revenue_in, cash_out, closing, runway], 2):
            cell = ws.cell(row=i, column=col, value=val)
            cell.number_format = USD_FMT if col != 7 else "0.0"
            cell.border = BORDER
        cash = closing

    ws["A38"] = "Form field hints (Step 6)"
    ws["A38"].font = Font(bold=True)
    ws["A39"] = "Funding raised till date"
    ws["B39"] = f"${prior:,.0f} pre-seed"
    ws["A40"] = "Monthly burn rate"
    ws["B40"] = f"${projections[0]['net_burn']:,.0f}/month (from model)"
    ws["A41"] = "Runway (months)"
    ws["B41"] = f"{cash / max(projections[-1]['net_burn'], 1):.0f} (post-raise, month 36)"


def _add_use_of_funds(ws, data: dict) -> None:
    ws.title = "Use of Funds"
    total = float(data.get("amount_raising", 500000))
    allocations = data.get(
        "use_of_funds_split",
        [
            ("Product & Engineering", 0.40, "Core ML infra, platform APIs"),
            ("Go-to-Market", 0.25, "Enterprise pilots, developer relations"),
            ("Cloud & Compute", 0.20, "GPU credits, distributed training"),
            ("Operations & Legal", 0.10, "Compliance, IP, admin"),
            ("Reserve / Buffer", 0.05, "12-month contingency"),
        ],
    )

    headers = ["Category", "% of Round", "Amount (USD)", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, 4)

    r = 2
    for name, pct, note in allocations:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=pct).number_format = PCT_FMT
        ws.cell(row=r, column=3, value=total * pct).number_format = USD_FMT
        ws.cell(row=r, column=4, value=note)
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=2, value=1).number_format = PCT_FMT
    ws.cell(row=r, column=3, value=total).number_format = USD_FMT

    ws["A10"] = "Paste into form Step 7 — Use of Funds"
    ws["A11"] = data.get(
        "use_of_funds_text",
        "40% product/engineering, 25% GTM, 20% cloud compute, 10% ops/legal, 5% reserve.",
    )


def _add_summary(ws, projections: list[dict], data: dict) -> None:
    ws.title = "Investor Summary"
    _set_col_widths(ws, {1: 28, 2: 16, 3: 16, 4: 16})

    def ysum(year: int, key: str) -> float:
        start = (year - 1) * 12
        end = year * 12
        return sum(p[key] for p in projections[start:end])

    headers = ["Metric", "Year 1", "Year 2", "Year 3"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, 4)

    metrics = [
        ("Revenue (USD)", "revenue"),
        ("Gross profit", "gross_profit"),
        ("EBITDA", "ebitda"),
        ("Ending MRR", "mrr"),
        ("Ending customers", "customers"),
    ]
    r = 2
    for label, key in metrics:
        ws.cell(row=r, column=1, value=label)
        for y in range(1, 4):
            if key == "mrr":
                val = projections[y * 12 - 1][key]
            elif key == "customers":
                val = projections[y * 12 - 1][key]
            else:
                val = ysum(y, key)
            cell = ws.cell(row=r, column=y + 1, value=val)
            cell.number_format = USD_FMT if key != "customers" else "0"
        r += 1

    _style_section(ws, r + 1, "COPY INTO FUNDING FORM", 4)
    r += 2
    y3_rev = ysum(3, "revenue") * 12  # approximate annualized from monthly sum * not quite - use last month arr
    last = projections[-1]
    hints = [
        ("Revenue projections (Step 6)", data.get("revenue_projections_text", "")),
        ("Current revenue", data.get("current_revenue_text", "")),
        ("Growth rate", data.get("growth_rate_text", "")),
        ("Number of customers", data.get("customers_text", "")),
        ("Amount raising", data.get("amount_raising_text", "")),
        ("Financial model link (Step 10)", data.get("financial_model_link", "See attached workbook / repo path")),
    ]
    for label, text in hints:
        ws.cell(row=r, column=1, value=label)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(row=r, column=2, value=text)
        r += 1

    ws.cell(row=r, column=1, value="Implied post-money (at ask)")
    equity = float(str(data.get("equity_pct", 10)).replace("%", "")) / 100
    post_money = float(data.get("amount_raising", 500000)) / equity if equity else 0
    ws.cell(row=r, column=2, value=post_money).number_format = USD_FMT


def build_workbook(data: dict) -> Workbook:
    wb = Workbook()
    projections = _project_revenue(data)
    _add_assumptions(wb.active, data)
    _add_monthly_pl(wb.create_sheet(), projections)
    _add_cashflow(wb.create_sheet(), projections, data)
    _add_use_of_funds(wb.create_sheet(), data)
    _add_summary(wb.create_sheet(), projections, data)
    return wb


APEX_DEMO = {
    "startup_name": "Apex AI Labs",
    "industry_sector": "AI / ML",
    "business_model": "SaaS",
    "current_stage": "MVP",
    "starting_mrr": 8000,
    "mom_growth": 0.12,
    "churn": 0.03,
    "starting_customers": 24,
    "arpu": 333,
    "prior_raise": 150000,
    "cash_on_hand": 180000,
    "current_burn": 42000,
    "amount_raising": 500000,
    "funding_stage": "Seed",
    "equity_offered": "10%",
    "equity_pct": 10,
    "cogs_pct": 0.38,
    "headcount_start": 6,
    "headcount_end": 14,
    "fte_cost": 8500,
    "round_close_month": 1,
    "use_of_funds_text": "40% hiring core engineers, 25% enterprise GTM, 20% cloud/GPU compute credits, 10% security & compliance, 5% reserve.",
    "current_revenue_text": "$8,000 MRR (~$96K ARR), usage-based inference API",
    "growth_rate_text": "12% MoM logo revenue growth; 3% monthly churn",
    "customers_text": "24 paying teams; 180 registered developers",
    "revenue_projections_text": "Y1: $250K ARR | Y2: $1.1M ARR | Y3: $3.8M ARR (base case)",
    "amount_raising_text": "$500,000 USD",
    "financial_model_link": "docs/financial-model/samples/Apex_AI_Labs_Financial_Model.xlsx",
}

TEMPLATE_DEFAULTS = {
    "startup_name": "[Your Startup]",
    "industry_sector": "AI / ML",
    "business_model": "SaaS",
    "current_stage": "MVP",
    "starting_mrr": 0,
    "mom_growth": 0.10,
    "churn": 0.04,
    "starting_customers": 0,
    "arpu": 0,
    "prior_raise": 0,
    "cash_on_hand": 50000,
    "current_burn": 25000,
    "amount_raising": 500000,
    "funding_stage": "Seed",
    "equity_offered": "10%",
    "equity_pct": 10,
    "revenue_projections_text": "Y1: $___ | Y2: $___ | Y3: $___",
    "current_revenue_text": "e.g. $X MRR or pre-revenue",
    "growth_rate_text": "e.g. 10% MoM",
    "customers_text": "e.g. 50 paying customers",
    "amount_raising_text": "$500,000 USD",
    "financial_model_link": "Upload file or Google Drive link in Step 10",
    "use_of_funds_text": "Describe allocation across product, GTM, hiring, infra.",
}


def main() -> None:
    os.makedirs(SAMPLES_DIR, exist_ok=True)

    template_path = os.path.join(OUT_DIR, "FSV_Financial_Model_Template.xlsx")
    sample_path = os.path.join(SAMPLES_DIR, "Apex_AI_Labs_Financial_Model.xlsx")

    build_workbook(TEMPLATE_DEFAULTS).save(template_path)
    build_workbook(APEX_DEMO).save(sample_path)

    print(f"Created: {template_path}")
    print(f"Created: {sample_path}")


if __name__ == "__main__":
    main()
